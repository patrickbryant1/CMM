#! /usr/bin/env python
# -*- coding: utf-8 -*-

'''
This is a program that takes an xl workbook, reads it and uses its
sheets to perform filtering as specified by the text file.

-Remember that you cannot write too many rows to an excel sheet,
there is a limitation!
'''

import sys
import os
import pdb
import xlrd
import argparse
import openpyxl

#Arguments for argparse module:
parser = argparse.ArgumentParser(description = '''
This is a program that takes an xl workbook, reads it and uses its
sheets to perform filtering as specified by text files.

-Remember that you cannot write too many rows to an excel sheet,
there is a limitation!
''')
 
parser.add_argument('workbook_r', nargs=1, type= str,
                  default=sys.stdin, help = 'path to excel file with variants to be opened')


parser.add_argument('ref_dbs', nargs=1, type= str,
                  default=sys.stdin, help = '''path to text file containing column positions
                  for ref_dbs to filter on.''')

parser.add_argument('zygosity_positions', nargs=1, type= str,
                  default=sys.stdin, help = '''path to text file containing
                  zygosity_positions to filter on.''')

parser.add_argument('out_name', nargs=1, type= str,
                  default=sys.stdin, help = '''str containing requested name of output file''')

###########################################################################################
#Functions

def text_to_list(file_name):
    '''A function that reads a text file and creates a list of its rows.
    Input: text_file (txt)
    Output: filter_options (list)
    '''

    filter_options = [] #empty list to store options in
    with open(file_name, 'r') as infile:
        for line in infile:
            line = line.rstrip('\n')
            filter_options.append(line)
            
    return filter_options

def encode_ascii(xl_sheet, row_idx, col_idx):
    '''A function that encodes and retrieves excel cell values
    Input: row and column positions
    Output: cell_value
    '''
    cell_value = str(xl_sheet.cell(row_idx, col_idx).value).encode('ascii','ignore')
    return cell_value

def get_zygosity_positions(zygosity_positions):
    '''A function that organizes zygosity positions according to families
    '''

    share_pos = [] #List to store colummn positions for those that should share
    not_share_pos = [] #List to store colummn positions for those that should not share
    share_families= [] #List to store found family numbers
    not_share_families = []

    #Get positions for zygosities
    for item in zygosity_positions:
        item = item.split(' ') #split on space
        family_number = item[0].split('-')[0] #Get family number
        
        if item[2] == '1': #If it is a share item
            if family_number in share_families:
                for i in range(0, len(share_pos)):
                    if share_pos[i][0].split('/')[0] == family_number:
                        share_pos[i].append(family_number+'/'+item[1]) 
                        break
            else:
                share_families.append(family_number)
                share_pos.append([family_number+'/'+item[1]])

        if item[2] == '0': #If it is a not share item
            if family_number in not_share_families:
                for i in range(0, len(not_share_pos)):
                    if not_share_pos[i][0].split('/')[0] == family_number:
                        not_share_pos[i].append(family_number+'/'+item[1]) 
                        break
            else:
                not_share_families.append(family_number)
                not_share_pos.append([family_number+'/'+item[1]])

    return(share_pos, not_share_pos)

def find_shared(share_pos, not_share_pos, sheet_r, row_idx, above_t):
    '''Finds the variants shared between the samples as specified by share_pos and not_share_pos
        Input = share_pos, not_share_pos, sheet_r, row_idx, above_t
        Output = shared (bool)
    '''
   
    
    shared_families = [] #Keep track of families that share
    shared = False #See if the variants are shared
   
    for family in share_pos:
        n_share = 0 #Keep track of how many that share
        
        for individual in family:
            pos = individual.split('/')[1] #Get column position
            zyg = encode_ascii(sheet_r, row_idx, int(pos)) #zygosity to match
            if zyg =='.':  #If the zygosity cannot be assessed, it is disregarded            
                break   
            else:
                if zyg == 'other': #If it is other, it is another variant
                    break
                if above_t == True:
                    if zyg == "wt" or not zyg or zyg == 'het': #If the zyg is empty, it is wt
                        n_share +=1
                else: #If above_t == False
                    if zyg == "hom" or zyg =='het':
                        n_share +=1
            
        if n_share == len(family):
            shared = True
            shared_families.append(individual.split('/')[0])
            #print encode_ascii(sheet_r, row_idx, 6)
            #print shared_families

    #Check the ones that should not share
    if shared == True and not_share_pos:
        found = False #Keep track of not_shared
        for family in not_share_pos:
            if family[0].split('/')[0] in shared_families:
                found = not_shared(found, family, above_t, sheet_r, row_idx)
                if found == True:
                    break

        if found == False: #If the not_share criteria are not fulfilled
            shared = False

    return(shared)
            
def not_shared(found, family, above_t, sheet_r, row_idx):
    '''Check if those that should not share do not
    '''

    
    n_share = 0 #Keep track of how many that do not share

    for individual in family:
        pos = individual.split('/')[1]
        zyg = encode_ascii(sheet_r, row_idx, int(pos)) #zygosity to match
        if zyg =='.':  #If the zygosity cannot be assessed, it is disregarded            
            break   
        if zyg == 'het': #shared
            break
        else:
            if zyg == 'other': #If it is other - that should be fine, right?! Can't exclude these
                n_share+=1
                
            if above_t == False:
                if zyg == "wt" or not zyg: #If the zyg is empty, it is wt
                    n_share +=1
            if above_t == True:
                if zyg == "hom":
                    n_share +=1


    if n_share == len(family):
        found = True
                
    return(found)

def filter_sheet(workbook_r, name, ref_dbs, zygosity_positions):        
    '''A function that takes an xl-sheet
    as input and finds the shared variants as specified in zygosity positions
    Input: workbook_r, name, ref_dbs, zygosity_positions
    Output: None
    '''
    #Order zygosity positions
    (share_pos, not_share_pos) = get_zygosity_positions(zygosity_positions)
    print share_pos
    print not_share_pos

    #Create an excel workbook and a sheet to write to
    workbook_w = openpyxl.Workbook()
    sheet_w = workbook_w.active #Get active sheet
    sheet_w.title = 'filtered'

    shared = False

    for num in range(0, 1):#workbook_r.nsheets):
        sheet_r = workbook_r.sheet_by_index(num) #Open sheet_num
	
        row_idx = 0 #Row to transfer from
        row_n = 1 #To keep track of which row to write to
        #transfer first row
        write_to_sheet(row_idx, sheet_w, sheet_r, row_n)

        

        for row_idx in range(1, sheet_r.nrows):
            above_t = False
            for ref_db in ref_dbs:
                ref_db = ref_db.split(' ') #split each list item on space
                MAF = encode_ascii(sheet_r, row_idx, int(ref_db[1]))#filter on reference databases MAFs
                if MAF and MAF != '.': #Checks if MAF is empty or NA
                    MAF = float(MAF)
                    if MAF > float(ref_db[3]): #Filtering on MAF above threshold
                        above_t = True
            
            shared = find_shared(share_pos, not_share_pos, sheet_r, row_idx, above_t)
            if shared == True:
                
                    row_n += 1
                    write_to_sheet(row_idx, sheet_w, sheet_r, row_n)
                    
            else:
                continue

    workbook_w.save('shared_'+name+'.xlsx')

    return None

def write_to_sheet(row_idx, sheet_w, sheet_r, row_n):
    '''A function that writes data into a sheet in the excel workbook.
    Input: row_idx, sheet_w, xl_sheet
    Output: None
    '''

    #Iterate over all columns
    for col_idx in range(0, sheet_r.ncols):
        cell = sheet_r.cell(row_idx, col_idx).value
        if type(cell) == float:
            	cell = str(cell) 
        cell = cell.encode('ascii','ignore')
        sheet_w.cell(row = row_n, column = col_idx+1).value = cell #Openpyxl is not zero indexed, but xlrd is

    return None
    
    
#Main program
args = parser.parse_args()

try:
    workbook_r = xlrd.open_workbook(args.workbook_r[0])
    #The opened workbook is closed automatically after use
    
    #Checks if file is empty
    if os.stat(args.workbook_r[0]).st_size == 0:
       print "Empty file."       
except IOError:
    print 'Cannot open', args.workbook_r[0]
else:


    #Arguments to filter on
    ref_dbs = text_to_list(args.ref_dbs[0])
    zygosity_positions = text_to_list(args.zygosity_positions[0])
    
    #Perform the filtering
    name = args.out_name[0]
                                            #name is a list here
    filter_sheet(workbook_r, name, ref_dbs, zygosity_positions)


    
