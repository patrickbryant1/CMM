#! /usr/bin/env python
# -*- coding: utf-8 -*-


import sys
import os
import pdb
import xlrd
import argparse
import openpyxl 
from mpl_toolkits.mplot3d import Axes3D
import matplotlib.pyplot as plt
import numpy as np
from matplotlib import cm

#Arguments for argparse module:
parser = argparse.ArgumentParser(description = '''A program that reads a file
containing haplotypes and aligns them at their correct genomic position
according to increasing window size.
Make sure all data from p-link is put in - otherwise the SNPs may
align incorrectly.''')
 
parser.add_argument('hap_file', nargs=1, type= str,
                  default=sys.stdin, help = 'path to file with haplotypes to be opened')

parser.add_argument('instructions', nargs=1, type= str,
                  default=sys.stdin, help = '''path to instructions containing criteria to filter haplotypes on:
                    start, end, p_val and OR.''')

###########################################################################################
#Functions
def text_to_list(file_name):
    '''A function that reads a text file and creates a list of its rows.
    Input: text_file (txt)
    Output: filter_options (list)
    '''

    filter_options = [] #empty list to store options in
    with open(file_name, 'r') as infile:
        for line in infile:
            line = line.rstrip('\n')
            filter_options.append(line)
            
    return filter_options

def hap_map(hap_file, name, instructions):        
    '''A function that takes an xl-sheet as input and aligns
    its haplotypes according to increasing window size.
    Input: workbook_r, name
    Output: None
    '''

    #Get instructions
    out_start = instructions[0].split(' ')[1]
    out_end = instructions[1].split(' ')[1]
    p_threshold = instructions[2].split(' ')[1] #The p-value threshold
    OR_threshold = instructions[3].split(' ')[1]
    
    #Create an excel workbook and a sheet to write to
    workbook_w = openpyxl.Workbook()
    sheet_w = workbook_w.active #Get active sheet
    sheet_w.title = 'Visualized'
   
    with open(hap_file, 'r') as infile:
        hap_lens = [] #List with haplotype lengths
        hap_lists = [] #List to store haplotype lists
        pos_to_rs = [] #Connect rs to pos

        line_number = 0 #Keep track of line being read
        for line in infile:
            if line_number >0:
                line = line.rstrip('\n')
                line = line.split()

                start = line[3] #Get start position of haplotype (first SNP)
                end = line[4] #Get end position of haplotype (second SNP)
                rs_1 = line[5]  #Get rs for start position
                rs_2 = line[6]  #Get rs for end position
                haplotype = line[7] #Get haplotype
                F = line[8]  #Get frequqncy of haplotype in samples
                OR = line[9]  #Get odds ratio
                p_val = line[11] #Get p-value
              

                start_rs = (start, rs_1) #Create tuple of start position and rs
               
                #Check output criteria are fulfilled'
                if p_val == 'NA':
                    continue
                if float(out_start) <= float(start) and float(end) <= float(out_end): 
                    if start_rs not in pos_to_rs:
                        pos_to_rs.append(start_rs) #Will only append unique
                    if float(p_val) < float(p_threshold) and float(OR) > float(OR_threshold): #If the variant meets the significance and OR
                        hap_key = start + '/' + haplotype + '/' + rs_1 + '/' + rs_2 + '/' + p_val + '/' + F + '/' + OR 
    
                        (hap_lists, hap_lens) = hap_sort(hap_key, haplotype, hap_lens, hap_lists)
                
            line_number+=1 #Increase line_number

    #Write headers
    write_headers(sheet_w)
    #Write rs
    pos_to_rs = write_rs(pos_to_rs, sheet_w)
    #Write haps
    transpose(hap_lists, sheet_w, hap_lens, pos_to_rs)

    workbook_w.save('transposed_'+name[0]+'.xlsx')

    return None

def encode_ascii(xl_sheet, row_idx, col_idx):
    '''A function that encodes and retrieves excel cell values
    Input: row and column positions
    Output: cell_value
    '''
    cell_value = str(xl_sheet.cell(row_idx, col_idx).value).encode('ascii','ignore')
    return cell_value
    
def hap_sort(hap_key, haplotype, hap_lens, hap_lists):
    '''A function that creates lists and adds haplotypes to
    these lists according to the haplotype sizes
    Input = hap_key(str), haplotype(str), hap_lens(list), hap_lists(list)
    Output po= hap_lists(list), hap_lens(list)
    '''
    size = str(len(haplotype)) 

    
    #If the haplotype length has not been found
    if size not in hap_lens:
        hap_lens.append(size) #Be sure strs and ints match
        hap_lists.append([hap_key]) #Add hap_key   
        
    #If the haplotype length has been found
    else:
        for hap_list in hap_lists:
            if len(hap_list[0].split('/')[1]) == len(haplotype): #Haplotypes of equal size should be in the same list
                    hap_list.append(hap_key) #Make sure hap_keys are unique since there can be
                    break                         #many different haplotypes with equal size
            else:
                continue  # only executed if the inner loop did NOT break
            break  # only executed if the inner loop DID break    
    
    return hap_lists, hap_lens

def write_headers(sheet_w):
    '''A function that writes the headers to the
    workbook
    '''
    sheet_w.cell(row=6,column=1).value = 'Position' 
    sheet_w.cell(row=6,column=2).value = 'SNP1'

    write_vertical('P-value', 1, 2, sheet_w)
    write_vertical('P-rank', 2, 2, sheet_w)
    write_vertical('OR', 3, 2, sheet_w)
    write_vertical('F', 4, 2, sheet_w)
    write_vertical('SNP2', 5, 2, sheet_w)

    return None

    
def take_first(item):
    '''Return first list element
    Input = item (list of lists)
    Output = item[0] (float)
    '''
    return float(item[0])
    
def write_rs(pos_to_rs, sheet_w):
    '''A function that writes positions and
    rs numbers to an excel sheet as well as
    indexing them
    '''
    
    #Sort on position
    pos_to_rs = sorted(pos_to_rs, key = take_first)
      
    for i in range(0,len(pos_to_rs)):
        #Write position
        item = pos_to_rs[i]
        cell = str(item[0]).encode('ascii','ignore')
        #sheet_w.write(i,0, cell) #Row i, column 0
        sheet_w.cell(row=i+7,column=1).value = cell #Row i, column 1
        #Write rs
        cell = str(item[1]).encode('ascii','ignore')
        sheet_w.cell(row=i+7,column=2).value = cell #Row i, column 2        
        pos_to_rs[i] = pos_to_rs[i] + (str(i+7),) #Add index to rs

    return pos_to_rs
        
def get_hap_len(item):
    '''A function that splits item on / and
    returns the haplotype length for sorting
    Input = item (list of lists)
    Output = length (int)
    '''
    
    length = len(item[0].split('/')[1])
    return length

def get_start(item):
    '''A function that splits item on / and
    returns the SNP start pos for sorting
    Input = item (list of lists)
    Output = start (float)
    '''
    
    start = float((item.split('/')[0]))
    return start
     
def p_sort(item):
   '''A function that splits item on / and
    returns the haplotype length for sorting
    Input = item (list of lists)
    Output = p_val
    '''
   p = float((item.split('/')[4]))
   return p



def transpose(hap_lists, sheet_w, hap_lens, pos_to_rs):
    '''A function that writes data into a sheet in the excel workbook.
    Input: hap_dicts, sheet_w
    Output: None
    '''
    #For surface plot
    X = []
    Y = []
    Z = []
    
    #Sort hap_lists on haplotype length
    hap_lists = sorted(hap_lists, key = get_hap_len)
    #Sort the contents of each hap_list on position
    for i in range(0, len(hap_lists)): #Remember to do a range when changing lists!
        hap_lists[i] = sorted(hap_lists[i], key = p_sort)#Sort on p-val and assign rank
        for j in range(0,len(hap_lists[i])):
            hap_lists[i][j]+='/'+str(float(j)+1)
        
        hap_lists[i] = sorted(hap_lists[i], key = get_start)#Sort on position of first SNP

    #When you print out the haps you want the ones of equal size in positional order
    col_idx = 2 #Keep track of column number
    for hap_list in hap_lists:
        for item in hap_list:
            item = item.split('/')
            haplotype = item[1]
            rs_1 = item[2]
            rs_2 = item[3]
            p_val = item[4]
            F = item[5]
            OR = item[6]
            p_rank = item[7]

            for pr in pos_to_rs:
                if rs_1 == pr[1]: #Match on rs
                    row_idx = int(pr[2])
                    col_idx += 1 #Write to next column
                    (X, Y, Z) = write_haps(haplotype, rs_1, rs_2, p_val, p_rank, F, OR, row_idx, col_idx, sheet_w, X, Y, Z)
                    break
                
                
    #surface_plot(X,Y,Z)
    return None

def write_vertical(item, i, j, sheet_w):
    '''A function that writes the contents of item
    to th cell row i,col j in sheet_w.
    '''
    cell = item.encode('ascii','ignore')
    sheet_w.cell(row=i, column=j).value= cell
    sheet_w.cell(row=i, column=j).alignment = openpyxl.styles.Alignment(text_rotation = 90)

    return None

def write_haps(haplotype,rs_1, rs_2, p_val, p_rank, F, OR, row_idx, col_idx, sheet_w, X, Y, Z):
    '''A function that writes the haplotype into an excel
    spread sheet.
    Input =
    Output = None
    '''
    
    #Write p-val
    write_vertical(p_val, 1, col_idx, sheet_w)
    #Write p-val rank
    write_vertical(p_rank, 2, col_idx, sheet_w)
    #Write OR
    write_vertical(OR, 3, col_idx, sheet_w)
    #Write F
    write_vertical(F, 4, col_idx, sheet_w)
    #Write rs_2
    write_vertical(rs_2, 5, col_idx, sheet_w)
    #Write rs_1
    write_vertical(rs_1, 6, col_idx, sheet_w)
    
    for base in haplotype:
        X.append(row_idx)
        Y.append(col_idx)
        Z.append(1/float(p_val))
        cell = base.encode('ascii','ignore')
        sheet_w.cell(row=row_idx,column=col_idx).value = cell
        row_idx +=1

    return X, Y, Z

def surface_plot(X,Y,Z):
    #Convert to numpy arrays
    X = np.array(X)
    Y = np.array(Y)
    Z = np.array(Z)
    # Plot the surface.
    fig = plt.figure()
    
    #ax = fig.gca(projection='3d')
    ax = Axes3D(fig)
    surf = ax.plot_trisurf(X, Y, Z, cmap=cm.coolwarm, linewidth=0, antialiased=False)
    fig.colorbar(surf, shrink=0.5, aspect=5)
    plt.show()
#Main program
     
args = parser.parse_args()

#Path to hap file
hap_file = (args.hap_file[0])

#Get criteria for output
instructions = text_to_list(args.instructions[0])
    
#Perform the alignment and ordering
name = str(args.hap_file[0]).split('/')[-1:] #name is only the last part of path
                                            #name is a list here
hap_map(hap_file, name, instructions)


    
