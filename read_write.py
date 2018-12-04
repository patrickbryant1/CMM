#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Created on Tue Nov 20 14:31:06 2018

@author: patbry
"""

import openpyxl
import pdb

def text_to_list(file_name):
    '''A function that reads a text file and creates a list of its rows.
    Input: text_file (txt)
    Output: filter_options (list)
    '''

    filter_options = [] #empty list to store options in
    with open(file_name, 'r') as infile:
        for line in infile:
            line = line.rstrip('\n')
            filter_options.append(line)
            
    return filter_options

def encode_ascii(xl_sheet, row_idx, col_idx):
    '''A function that encodes and retrieves excel cell values
    Input: row and column positions
    Output: cell_value
    '''
    cell_value = str(xl_sheet.cell(row_idx, col_idx).value).encode('ascii','ignore')
    return cell_value

def write_to_sheet(row_r, sheet_w, sheet_r, row_w):
    '''A function that writes data into a sheet in the excel workbook.
    Input: row_idx, sheet_w, xl_sheet
    Output: None
    '''

    #Iterate over all columns
    for col_r in range(0, sheet_r.ncols):
        cell = sheet_r.cell(row_r, col_r).value
        if type(cell) == float:
            	cell = str(cell) 

        cell = cell.encode('ascii','ignore')
        sheet_w.cell(row = row_w, column = col_r+1).value = cell #Openpyxl is not zero indexed, but xlrd is not

    return None

def fill(sheet_w, row_n, row_idx, col_idx, color_code):
    '''Fill cell with color according to color_code
    Output = None
    '''

    color = openpyxl.styles.colors.Color(rgb=color_code)
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=color)
    sheet_w.cell(row = row_n, column = col_idx+1).fill = my_fill



